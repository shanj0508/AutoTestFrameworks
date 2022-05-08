'''
基于Excel文件的内容去进行读取，并结合获取的数据进行自动化的测试执行
'''

import openpyxl

# 读取excel中的用例正文

from ui_test_frame_excel_driver.ui_keyword.ui_keyword import WebKey


def read(file):
    excel = openpyxl.load_workbook(file)
    for name in excel.sheetnames:
        sheet = excel[name]
        print('***********{}**************'.format(name))
        web_key = None
        for values in sheet.values:
            # 如果第一个单元格是int类型，则表示进入了测试用例的正文
            if type(values[0]) is int:
                # print(values)
                # 接收每一行操作行为对应的参数内容
                data = {}
                data['name'] = values[2]
                data['value'] = values[3]
                data['content'] = values[4]
                data['expect'] = values[6]

                print(data)
                # 清除参数字典中为None的参数键值对
                for key in list(data.keys()):
                    if data[key] is None:
                        del data[key]
                print(data)

                # 基于操作行为和对应参数来执行自动化操作
                '''
                    用例的操作行为主要分为：
                        1. 实例化，通过一个操作行为实例化关键字驱动类对象
                        2. 常规操作，通过调用已实例化的对象，执行对应的函数。
                        3. 断言操作，判断预期与实际是否符合，将结果填入测试用例中。
                '''
                # 实例化关键字驱动
                if values[1] == 'open_browser':
                    web_key = WebKey(data["content"])
                    # driver = webdriver.Chrome()
                # 断言
                elif 'assert' in values[1]:
                    status = getattr(web_key, values[1])(**data)
                    # 基于断言结果True or False来进行写入的操作
                    if status:
                        sheet.cell(row=values[0] + 2, column=8).value = 'Pass'
                    else:
                        sheet.cell(row=values[0] + 2, column=8).value = 'Failed'
                # 常规操作
                else:
                    getattr(web_key, values[1])(**data)
    excel.save(file)
