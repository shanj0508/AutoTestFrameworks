import openpyxl

# 读取excel文件
from api_test_frame_excel_driver.api_keyword.api_keyword import ApiKey


def read(file):
    excel = openpyxl.load_workbook(file)
    ak = ApiKey()
    # 读取所有的sheet页中全部单元格内容

    # 1.获取excel中所有的sheet页名称，返回list格式的内容
    sheets = excel.sheetnames
    print(sheets)
    # 2.通过循环读取sheets中的内容，以key的形式赋值给excel
    for sheet in sheets:
        print(sheet)
        print('*' * 20)

        # 3.每一个sheet中的单元格内容输出
        for row_index, value in enumerate(excel[sheet].values):
            # 逐行读取excel用例内容，实现文件驱动自动化执行
            if type(value[0]) is int:  # 跳过首行
                # print(value[0])
                # 准备测试数据
                """
                        01 用例编号
                        02 用例标题
                        03 测试地址
                        04 请求路径
                        05 请求方法
                        06 请求头
                        07 请求参数
                        08 参数类型
                        09 校验字段
                        10 预期结果
                        11 实际结果
                        12 测试结果
                """
                url = value[2] + value[3]
                if value[5]:
                    headers = eval(value[5])
                if value[6]:
                    data = value[6]
                assert_attribute = value[8]
                expect_result = value[9]

                # 存在请求头
                if value[5]:
                    # 存在请求参数
                    if value[6]:
                        dict_data = {
                            "url": url,
                            "headers": headers,
                            value[7]: eval(data)
                        }


                    # 不存在请求参数
                    else:
                        dict_data = {
                            "url": url,
                            "headers": headers
                        }

                # 不存在请求头
                else:
                    # 存在请求参数
                    if value[6]:
                        dict_data = {
                            "url": url,
                            value[7]: eval(data)
                        }

                    # 不存在请求参数
                    else:
                        dict_data = {
                            "url": url,
                        }

                print(dict_data)

                # 模拟请求
                # getattr(object, name[, default])
                res = getattr(ak, value[4])(**dict_data)
                print("res", res.status_code)

                try:
                    # 结果校验
                    result = ak.get_text(res.text, assert_attribute)
                    print("==========实际结果=========")
                    print(result)
                    print("===================")
                    print(result == expect_result)
                    excel[sheet].cell(row=row_index + 1, column=11).value = result
                    if result == expect_result:
                        excel[sheet].cell(row=row_index + 1, column=12).value = "Pass"
                    else:
                        excel[sheet].cell(row=row_index + 1, column=12).value = "Failed"

                except Exception as e:
                    print("==========e=========")
                    print(e)
                    print("请求参数有误，请检查")
                    excel[sheet].cell(row=row_index + 1, column=11).value = res.status_code
                    excel[sheet].cell(row=row_index + 1, column=12).value = "Failed"
                finally:
                    excel.save(file)
